from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.db.models.functions import Now
from django.test import TestCase
from openpyxl import Workbook, load_workbook

import rm
from rm.constants import CONTRACTEN, NEGOMETRIX, RowStatus, FileStatus
from rm.interface_file_util import check_file_and_interface_type
from rm.models import System, DataSetType, InterfaceDefinition, InterfaceCall, Mapping, DataPerOrgUnit, Contract
from rm.negometrix import NegometrixInterfaceFile, handle_negometrix_file_row
from rm.test_util import set_up_user_with_interface_call_and_contract, create_superuser
from rm.views import process_file
from users.models import OrganizationalUnit


class NegometrixFindInterfaceDefinitionTests(TestCase):

    def test_get_interface_definition_fail_Negometrix_not_found(self):
        # static data (wrong)
        system = System.objects.create(name="FOUTGESPELD")
        data_set_type = DataSetType.objects.create(name=CONTRACTEN)
        interface_definition = InterfaceDefinition.objects.create(system=system,
                                                                  data_set_type=data_set_type,
                                                                  interface_type=InterfaceDefinition.UPLOAD)

        negometrix_file = NegometrixInterfaceFile("testfile.xlsx")
        with self.assertRaises(rm.models.System.DoesNotExist):
            interface_definition = negometrix_file.get_interface_definition()

    def test_get_interface_definition_fail_Foute_dataset(self):
        # static data (wrong)
        system = System.objects.create(name=NEGOMETRIX)
        data_set_type = DataSetType.objects.create(name="Foute Dataset")
        interface_definition = InterfaceDefinition.objects.create(system=system,
                                                                  data_set_type=data_set_type,
                                                                  interface_type=InterfaceDefinition.UPLOAD)

        negometrix_file = NegometrixInterfaceFile("testfile.xlsx")
        with self.assertRaises(rm.models.DataSetType.DoesNotExist):
            interface_definition = negometrix_file.get_interface_definition()


    def test_get_interface_definition_fail_no_interface_definition_forgot_UPLOAD_in_GET(self):
        system = System.objects.create(name=NEGOMETRIX)
        data_set_type = DataSetType.objects.create(name=CONTRACTEN)
        interface_definition = InterfaceDefinition.objects.create(system=system,
                                                                  data_set_type=data_set_type)
        negometrix_file = NegometrixInterfaceFile("testfile.xlsx")
        with self.assertRaises(rm.models.InterfaceDefinition.DoesNotExist):
            interface_definition = negometrix_file.get_interface_definition()

    def test_get_interface_definition_happy(self):
        system = System.objects.create(name=NEGOMETRIX)
        data_set_type = DataSetType.objects.create(name=CONTRACTEN)
        interface_definition = InterfaceDefinition.objects.create(system=system,
                                                                  data_set_type=data_set_type,
                                                                  interface_type=InterfaceDefinition.UPLOAD)
        negometrix_file = NegometrixInterfaceFile("testfile.xlsx")
        interface_definition = negometrix_file.get_interface_definition()
        self.assertIsInstance(interface_definition, InterfaceDefinition)


class NegometrixFileTests(TestCase):

    def setUp(self):

        # Create Superuser and Login
        self.user = create_superuser()
        self.client.force_login(self.user)

        # STATIC TOTAL_DATA_ROWS_RECEIVED
        self.system, created = System.objects.get_or_create(name=NEGOMETRIX)
        self.data_set_type, created = DataSetType.objects.get_or_create(name=CONTRACTEN)
        self.interface_definition, created = InterfaceDefinition.objects.get_or_create(
                                                                            system=self.system,
                                                                            data_set_type=self.data_set_type,
                                                                            interface_type=InterfaceDefinition.UPLOAD)
        self.org_unit, created = OrganizationalUnit.objects.get_or_create(name="PT: IaaS",
                                                                          type=OrganizationalUnit.TEAM)


    def test_check_valid_negometrix_excel_file(self):
        interfaceCall = InterfaceCall.objects.create(date_time_creation=Now(),
                                                     status='TestStatus',
                                                     filename='test_register_contract.xlsx',
                                                     interface_definition=self.interface_definition)
        file = "rm/test/resources/test_register_contract.xlsx"
        excelInterfaceFile = check_file_and_interface_type(file)
        self.assertTrue(isinstance(excelInterfaceFile, NegometrixInterfaceFile))

    def test_upload_valid_negometrix_excel_file_2_valid_rows(self):

        Mapping.objects.create(system=self.system,
                               org_unit=self.org_unit,
                               name="NPO/Technology/IAAS")

        interfaceCall = InterfaceCall.objects.create(
                        date_time_creation=Now(),
                        status='TestStatus',
                        filename='test_upload_valid_negometrix_excel_file_2_valid_rows.xlsx',
                        interface_definition=self.interface_definition)

        file = "rm/test/resources/test_upload_valid_negometrix_excel_file_2_valid_rows.xlsx"
        excelInterfaceFile = check_file_and_interface_type(file)

        self.assertTrue(isinstance(excelInterfaceFile, NegometrixInterfaceFile))

        excelInterfaceFile.process(interfaceCall)

        raw_data_set = interfaceCall.rawdata_set.all()
        self.assertEqual(len(raw_data_set), 3)
        errors = 0
        for raw_data in raw_data_set:
            if raw_data.status == RowStatus.DATA_ERROR:
                errors += 1
        self.assertEqual(errors, 0)

        #TODO make this work, InterfaceCall.number_of_rows etc
        # self.assertEqual(interface_call.number_of_rows_received, 3)

        contracten = interfaceCall.contracts()
        self.assertEqual(len(contracten), 2)

        contract1 = contracten[0]
        self.assertEqual(contract1.contract_nr, '44335')
        contract2 = contracten[1]
        self.assertEqual(contract2.contract_nr, '44336')

    def test_upload_valid_negometrix_excel_file_2_valid_rows_1_invalid_row(self):
        categorie = "NPO/Technology/IAAS"
        Mapping.objects.create(system=self.system, org_unit=self.org_unit, name=categorie)

        interfaceCall = InterfaceCall.objects.create(date_time_creation=Now(),
                                                     status='TestStatus',
                                                     filename='test_upload_valid_negometrix_excel_file_2_valid_rows_1_invalid_row.xlsx',
                                                     interface_definition=self.interface_definition)
        file = "rm/test/resources/test_upload_valid_negometrix_excel_file_2_valid_rows_1_invalid_row.xlsx"
        excelInterfaceFile = check_file_and_interface_type(file)
        self.assertTrue(isinstance(excelInterfaceFile, NegometrixInterfaceFile))

        excelInterfaceFile.process(interfaceCall)

        contracten = interfaceCall.contracts()
        self.assertEqual(len(contracten), 2)

        contract1 = contracten[0]
        self.assertEqual(contract1.contract_nr, '44335')
        contract2 = contracten[1]
        self.assertEqual(contract2.contract_nr, '44337')

        rawdata = interfaceCall.rawdata_set.all()
        self.assertEqual(len(rawdata), 4)
        self.assertEqual(rawdata[0].status, RowStatus.HEADER_ROW.name)
        self.assertEqual(rawdata[1].status, RowStatus.DATA_OK.name)
        self.assertEqual(rawdata[2].status, RowStatus.DATA_ERROR.name)
        self.assertEqual(rawdata[3].status, RowStatus.DATA_OK.name)


    def test_upload_valid_negometrix_file_and_see_contracts(self):

        Mapping.objects.create(system=self.system, org_unit=self.org_unit, name="NPO/Technology/IAAS")

        interfaceCall = InterfaceCall.objects.create(
                        date_time_creation=Now(),
                        status='TestStatus',
                        filename='test_upload_valid_negometrix_excel_file_2_valid_rows.xlsx',
                        interface_definition=self.interface_definition)

        file = "rm/test/resources/test_upload_valid_negometrix_excel_file_2_valid_rows.xlsx"
        excelInterfaceFile = check_file_and_interface_type(file)
        excelInterfaceFile.process(interfaceCall)

        response = self.client.get(f'/interfacecall/{interfaceCall.pk}/')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Data verversen details')
        self.assertContains(response, '44336') # contract nr


class NegometrixCountTests(TestCase):

    def setUp(self):

        # Create Superuser and Login
        self.user = create_superuser()
        self.client.force_login(self.user)

        # STATIC TOTAL_DATA_ROWS_RECEIVED
        self.system, created = System.objects.get_or_create(name=NEGOMETRIX)
        self.data_set_type, created = DataSetType.objects.get_or_create(name=CONTRACTEN)
        self.interface_definition, created = InterfaceDefinition.objects.get_or_create(
                                                                            system=self.system,
                                                                            data_set_type=self.data_set_type,
                                                                            interface_type=InterfaceDefinition.UPLOAD)
        self.org_unit, created = OrganizationalUnit.objects.get_or_create(name="PT: IaaS",
                                                                          type=OrganizationalUnit.TEAM)

    def test_total_empty_header_data_recevied_error_and_ok(self):
        Mapping.objects.create(system=self.system, org_unit=self.org_unit, name="NPO/Technology/IAAS")

        interfaceCall = InterfaceCall.objects.create(
                        date_time_creation=Now(),
                        status='TestStatus',
                        filename='test_total_data_error_empty_rows.xlsx',
                        interface_definition=self.interface_definition)

        file = "rm/test/resources/test_total_data_error_empty_rows.xlsx"
        excelInterfaceFile = check_file_and_interface_type(file)
        excelInterfaceFile.process(interfaceCall)

        self.assertEqual(interfaceCall.number_of_rows_received, 4)
        self.assertEqual(interfaceCall.number_of_empty_rows, 1)
        self.assertEqual(interfaceCall.number_of_header_rows, 1)
        self.assertEqual(interfaceCall.number_of_data_rows_received, 2)
        self.assertEqual(interfaceCall.number_of_data_rows_error, 1)
        self.assertEqual(interfaceCall.number_of_data_rows_ok, 1)

        #TODO Test IGNORED - when we check on organizational department of the user


class FileWithContractTests(TestCase):

    def setUp(self):
        set_up_user_with_interface_call_and_contract(self)

    def test_register_contract(self):
        row_nr = 4
        file = open("rm/test/resources/test_register_contract.xlsx", "rb")
        workbook: Workbook = load_workbook(file)
        sheet = workbook.active
        count = 1
        values_row_4 = []
        for row_values in sheet.iter_rows(min_row=1,
                                          min_col=1,
                                          values_only=True):
            if count == 4:
                values_row_4 = row_values
                break
            count += 1

        interface_call = InterfaceCall.objects.create(date_time_creation=Now(),
                                                      status='TestStatus',
                                                      filename='test_register_contract.xlsx',
                                                      interface_definition=self.interface_definition)

        fields_with_position = dict(database_nr=0,
                                    contract_nr=1,
                                    contract_status=2,
                                    description=3,
                                    )

        mandatory_field_positions = (1, 2)
        mandatory_fields = ('a_field', 'another_field')

        status, msg = handle_negometrix_file_row(row_nr,
                                                 values_row_4,
                                                 interface_call,
                                                 fields_with_position,
                                                 mandatory_field_positions)

        expected_status = RowStatus.DATA_ERROR

        self.assertEqual(status, expected_status)


class NegometrixFileUploadTests(TestCase):

    def setUp(self):
        set_up_user_with_interface_call_and_contract(self, superuser=True)

    def test_upload_a_valid_excel_file(self):
        system_negometrix, created = System.objects.get_or_create(name=NEGOMETRIX)
        data_set_type_contracten, created = DataSetType.objects.get_or_create(name=CONTRACTEN)
        InterfaceDefinition.objects.get_or_create(name="Contracten upload",
                                                  data_set_type=data_set_type_contracten,
                                                  system=system_negometrix,
                                                  interface_type=InterfaceDefinition.UPLOAD)
        nr_int_calls_before = len(InterfaceCall.objects.all())
        file = open("rm/test/resources/a_valid_excel_file.xlsx", "rb")

        status, msg = process_file(file, self.user)
        nr_int_calls_after = len(InterfaceCall.objects.all())
        self.assertEqual(nr_int_calls_after, nr_int_calls_before + 1)

        interface_call: InterfaceCall = InterfaceCall.objects.last()
        self.assertEqual(interface_call.filename, "rm/test/resources/a_valid_excel_file.xlsx")
        self.assertEqual(interface_call.status, FileStatus.OK.name, msg=interface_call.message)
        self.assertEqual(interface_call.message, "")

    def test_upload_a_valid_excel_file(self):
        system_negometrix, created = System.objects.get_or_create(name=NEGOMETRIX)
        data_set_type_contracten, created = DataSetType.objects.get_or_create(name=CONTRACTEN)
        InterfaceDefinition.objects.get_or_create(name="Contracten upload",
                                                  data_set_type=data_set_type_contracten,
                                                  system=system_negometrix,
                                                  interface_type=InterfaceDefinition.UPLOAD)
        nr_int_calls_before = len(InterfaceCall.objects.all())
        file = open("rm/test/resources/a_valid_excel_file.xlsx", "rb")

        status, msg = process_file(file, self.user)
        nr_int_calls_after = len(InterfaceCall.objects.all())
        self.assertEqual(nr_int_calls_after, nr_int_calls_before + 1)

        interface_call: InterfaceCall = InterfaceCall.objects.last()
        self.assertEqual(interface_call.filename, "rm/test/resources/a_valid_excel_file.xlsx")
        self.assertEqual(interface_call.status, FileStatus.OK.name, msg=interface_call.message)
        self.assertEqual(interface_call.message, "")